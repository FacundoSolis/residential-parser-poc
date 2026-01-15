import openpyxl

excel_path = 'data/DE PAZ FRANCO QUINTILIANA_Checks.xlsx'
wb = openpyxl.load_workbook(excel_path)

print("=== SHEETS ===")
print(wb.sheetnames)
print()

# Ver primera sheet
ws = wb[wb.sheetnames[0]]

print("=== FIRST 20 ROWS ===")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 20:
        break
    print(f"Row {i}: {row}")
