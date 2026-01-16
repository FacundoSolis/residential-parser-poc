import openpyxl
from pathlib import Path

calculo_files = [
    'data/sample_1/Calculo.xlsx',
    'data/DE PAZ FRANCO QUINTILIANA/E1-4 OTROS DOCUMENTOS JUSTIFICATIVOS/CALCULO.xlsx',
]

for path in calculo_files:
    if not Path(path).exists():
        print(f"❌ Not found: {path}")
        continue
        
    print(f"\n{'='*80}")
    print(f"📊 {path}")
    print('='*80)
    
    wb = openpyxl.load_workbook(path)
    print(f"\nSheets: {wb.sheetnames}")
    
    # Ve primera sheet
    ws = wb[wb.sheetnames[0]]
    print(f"\nFirst 15 rows of '{wb.sheetnames[0]}':")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 15:
            break
        print(f"Row {i}: {row}")
