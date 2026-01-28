"""
Matrix Generator - Combines data from all document parsers into Excel output
UPDATED: Matches Travis's Excel template structure
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
from typing import Dict, Any, Optional, Tuple
import re
import os
import fitz


from .parsers.contrato_parser import ContratoParser
from .parsers.certificado_parser import CertificadoParser
from .parsers.factura_parser import FacturaParser
from .parsers.declaracion_parser import DeclaracionParser
from .parsers.cee_parser import CeeParser
from .parsers.registro_parser import RegistroParser
from .parsers.dni_parser import DniParser
from .parsers.calculo_parser import CalculoParser
from .parsers.ficha_parser import FichaParser
from openpyxl.drawing.image import Image as XLImage



class MatrixGenerator:
    """Generates correspondence matrix Excel from parsed documents"""
    
    PARSERS = {
        'CONTRATO': ContratoParser,
        'CERTIFICADO': CertificadoParser,
        'FACTURA': FacturaParser,
        'DECLARACION': DeclaracionParser,
        'CEE': CeeParser,
        'REGISTRO': RegistroParser,
        'DNI': DniParser,
        'CALCULO': CalculoParser,
        'FICHA': FichaParser,
    }
    
    DOCUMENT_PATTERNS = {
        'CONTRATO': [
            (r'contrato.*cesi[oó]n', None),
            (r'cesi[oó]n.*ahorro', None),
            (r'E0?4[-_\s]?1[-_\s]?1', None),
            (r'contrato', None),
            (r'convenio.*cae', None),
        ],
        'CEE': [
            (r'E0?1[-_\s]?3[-_\s]?6(?:[-_\s]?1)?', None),
            (r'cee.*final', None),
            (r'E0?4[-_\s]?3[-_\s]?6[-_\s]?1', None),
            (r'certificado.*eficiencia', None),
            (r'certificado.*energ[eé]tico', None),
            (r'certificaci[oó]n.*energ[eé]tica', None),

        ],
        'REGISTRO': [
            (r'E0?4[-_\s]?3[-_\s]?6[-_\s]?2', None),
            (r'registro\s*cee', None),
            (r'registro', r'mercantil'),
        ],

        'CERTIFICADO': [
            (r'certificado.*instalador', None),
            (r'E0?4[-_\s]?3[-_\s]?5', None),
            (r'cert.*instalador', None),
        ],
        'FACTURA': [
            (r'factura', None),
            (r'E0?4[-_\s]?3[-_\s]?3', None),
        ],
        'DECLARACION': [
            (r'declaraci[oó]n', None),
            (r'E0?4[-_\s]?3[-_\s]?2', None),
        ],
        'DNI': [
            (r'dni', None),
            (r'E0?4[-_\s]?4[-_\s]?1', None),
        ],
        'CALCULO': [
            (r'calculo', None),
            (r'c[aá]lculo', None),
            (r'ui.*rtotal', None),
            (r'E0?4[-_\s]?4[-_\s]?2', None),
        ],
        'FICHA': [
            (r'ficha.*res', None),
            (r'E0?4[-_\s]?3[-_\s]?1', None),
            (r'ficha', None),
        ],
        'INFORME': [
            (r'informe.*fotogr[aá]fico', None),
            (r'E0?4[-_\s]?3[-_\s]?4', None),
            (r'fotografico', None),
        ],
    }
    
    def __init__(self, folder_path: str):
        self.folder_path = Path(folder_path)
        self.parsed_data = {}
        self.file_mapping = {}
        
    def identify_document_type(self, filepath: Path) -> Optional[str]:
        """Identify document type using flexible pattern matching."""
        filename = filepath.stem
        parent_name = filepath.parent.name if filepath.parent != self.folder_path else ""
        search_text = f"{parent_name} {filename}".upper()
        full_path_text = str(filepath).upper()
        
        priority_order = ['CEE', 'CERTIFICADO', 'CONTRATO', 'REGISTRO', 'FACTURA', 
                         'DECLARACION', 'DNI', 'CALCULO', 'FICHA', 'INFORME']
        
        for doc_type in priority_order:
            patterns = self.DOCUMENT_PATTERNS.get(doc_type, [])
            for pattern, exclude_pattern in patterns:
                if re.search(pattern, search_text, re.IGNORECASE) or \
                   re.search(pattern, full_path_text, re.IGNORECASE):
                    if exclude_pattern and re.search(exclude_pattern, search_text, re.IGNORECASE):
                        continue
                    return doc_type
        
        return None
    
    def parse_all_documents(self):
        """Parse all documents in folder recursively"""
        print(f"\n📄 Parsing documents from: {self.folder_path}")
        
        supported_extensions = {'.pdf', '.jpg', '.jpeg', '.png'}
        excel_extensions = {'.xlsx', '.xls'}
        
        found_docs = {}
        all_files = list(self.folder_path.rglob("*"))
        
        for filepath in all_files:
            if not filepath.is_file():
                continue
                
            ext = filepath.suffix.lower()
            
            if ext in supported_extensions:
                doc_type = self.identify_document_type(filepath)
                if doc_type and doc_type in self.PARSERS:
                    if doc_type in found_docs:
                        existing = found_docs[doc_type]
                        if doc_type.lower() in filepath.name.lower() and \
                           doc_type.lower() not in existing.name.lower():
                            found_docs[doc_type] = filepath
                    else:
                        found_docs[doc_type] = filepath
                        
            elif ext in excel_extensions:
                if 'Checks' in filepath.name or 'output' in str(filepath):
                    continue
                doc_type = self.identify_document_type(filepath)
                if doc_type == 'CALCULO':
                    found_docs['CALCULO'] = filepath
        
        print(f"\n📋 Found {len(found_docs)} documents:")
        for doc_type, filepath in found_docs.items():
            print(f"   {doc_type}: {filepath.name}")
        
        print(f"\n🔍 Parsing...")
        # Procesa primero CONTRATO y FACTURA, luego el resto, DECLARACION después
        ordered_types = []
        for t in ["CONTRATO", "FACTURA"]:
            if t in found_docs:
                ordered_types.append(t)
        for t in found_docs:
            if t not in ordered_types and t != "DECLARACION":
                ordered_types.append(t)
        if "DECLARACION" in found_docs:
            ordered_types.append("DECLARACION")

        for doc_type in ordered_types:
            filepath = found_docs[doc_type]
            self.file_mapping[doc_type] = filepath

            if doc_type not in self.PARSERS:
                continue

            parser_class = self.PARSERS[doc_type]

            try:
                if doc_type == 'DECLARACION':
                    parser = parser_class(str(filepath))
                    # Pasa el contexto de todos los datos parseados hasta ahora
                    parser.context_data = dict(self.parsed_data)
                else:
                    parser = parser_class(str(filepath))
                data = parser.parse()
                self.parsed_data[doc_type] = data
                print(f"   ✓ Parsed {doc_type}")
            except Exception as e:
                print(f"   ✗ Error parsing {doc_type}: {e}")
                self.parsed_data[doc_type] = {
                    'document_type': doc_type,
                    'error': str(e)
                }
    
    def generate_excel(self, output_path: str, project_name: str = "Project"):
        """Generate Excel file with correspondence matrix - matches Travis's template"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Checks"
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[col].width = 30
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        section_font = Font(bold=True)
        
        # Headers row 3
        headers = {
            'B3': 'Info',
            'C3': 'E1-1-1\nCONTRATO CESION AHORROS',
            'D3': 'E1-3-1\nFICHA RES020',
            'E3': 'E1-3-2\nDECLARACION RESPONSABLE',
            'F3': 'E1-3-3\nFACTURA',
            'G3': 'E1-3-4\nINFORME FOTOGRAFICO',
            'H3': 'E1-3-5\nCERTIFICADO INSTALADOR',
            'I3': 'E1-3-6-1\nCEE FINAL',
            'J3': 'E1-3-6-2\nREGISTRO CEE',
            'K3': 'E1-4-1\nDNI',
            'L3': 'E1-4-2\nCALCULO UI RTOTAL',
        }
        
        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[3].height = 40
        
        current_row = 4
        
        # ==================== HOME OWNER SECTION ====================
        ws[f'A{current_row}'] = 'HOME OWNER'
        ws[f'A{current_row}'].font = section_font
        

        # Name
        ws[f'B{current_row}'] = 'Name'
        contr_name = self._get_value('CONTRATO', 'homeowner_name')
        decl_name = self._get_value('DECLARACION', 'homeowner_name')
        fact_name  = self._get_value('FACTURA', 'homeowner_name')
        dni_name   = self._get_value('DNI', 'name')
        calc_name  = self._get_value('CALCULO', 'client_name')

        # Fallback para columna K (DNI): si no hay nombre en DNI, usa DECLARACION o FACTURA
        best_name = ''
        for n in [dni_name, decl_name, fact_name, contr_name]:
            if n and n != 'NOT FOUND':
                best_name = n
                break

        ws[f'C{current_row}'] = contr_name if contr_name else best_name
        ws[f'E{current_row}'] = decl_name
        ws[f'F{current_row}'] = fact_name
        ws[f'K{current_row}'] = best_name
        ws[f'L{current_row}'] = calc_name
        current_row += 1

        # DNI
        ws[f'B{current_row}'] = 'DNI'
        contr_dni = self._get_value('CONTRATO', 'homeowner_dni')
        decl_dni  = self._get_value('DECLARACION', 'homeowner_dni')
        fact_dni  = self._get_value('FACTURA', 'homeowner_dni')
        dni_num   = self._get_value('DNI', 'dni_number')

        # Fallback para columna K (DNI): si no hay en DNI, usa DECLARACION o FACTURA
        best_dni = ''
        for d in [dni_num, decl_dni, fact_dni, contr_dni]:
            if d and d != 'NOT FOUND':
                best_dni = d
                break

        ws[f'C{current_row}'] = contr_dni
        ws[f'E{current_row}'] = decl_dni
        ws[f'F{current_row}'] = fact_dni
        ws[f'K{current_row}'] = best_dni
        current_row += 1

        
        # Address
        ws[f'B{current_row}'] = 'Address'

        contr_addr = self._get_value('CONTRATO', 'homeowner_address')
        contr_loc  = self._get_value('CONTRATO', 'location')  # a veces viene bien
        decl_addr = self._get_value('DECLARACION', 'homeowner_address')
        fact_addr  = self._get_value('FACTURA', 'homeowner_address')
        cert_addr  = self._get_value('CERTIFICADO', 'address')

        # Si decl_addr es vacío o NOT FOUND, usa contr_addr
        decl_addr_excel = decl_addr
        if not self._is_valid_cell_value(decl_addr, min_len=5):
            if self._is_valid_cell_value(contr_addr, min_len=5):
                decl_addr_excel = contr_addr

        ws[f'C{current_row}'] = contr_addr
        ws[f'E{current_row}'] = decl_addr_excel
        ws[f'F{current_row}'] = fact_addr

        current_row += 1

        
        # ==================== ACT SECTION ====================
        ws[f'A{current_row}'] = 'ACT'
        ws[f'A{current_row}'].font = section_font
        
        # Code
        ws[f'B{current_row}'] = 'Code (010/020)'

        c_code = self.to_010_020(self._get_value('CONTRATO', 'act_code'))
        ws[f'C{current_row}'] = c_code
        ws[f'D{current_row}'] = self.to_010_020(self._get_value('FICHA', 'act_code'))  
        ws[f'E{current_row}'] = self.to_010_020(self._get_value('DECLARACION', 'act_code'))
        ws[f'H{current_row}'] = self.to_010_020(self._get_value('CERTIFICADO', 'act_code'))
        ws[f'L{current_row}'] = self.to_010_020(self._get_value('CALCULO', 'act_code'))

        current_row += 1

        
        # Energy savings
        ws[f'B{current_row}'] = 'Energy savings (kWh)'

        contr_ae = self._get_value('CONTRATO', 'energy_savings')
        cert_ae  = self._get_value('CERTIFICADO', 'energy_savings')
        calc_ae  = self._get_value('CALCULO', 'ae')
        ws[f'D{current_row}'] = self._norm_num(cert_ae)



        # Columna C = CONTRATO
        ws[f'C{current_row}'] = self._norm_num(contr_ae)

        ws[f'H{current_row}'] = self._norm_num(cert_ae)
        ws[f'L{current_row}'] = self._norm_num(calc_ae)

        current_row += 1

        # Start date
        ws[f'B{current_row}'] = 'Start date'

        start = self._get_value('CERTIFICADO', 'start_date')
        contrato_start = self._get_value('CONTRATO', 'start_date')
        if not contrato_start:
            contrato_start = start
        ws[f'C{current_row}'] = contrato_start
        ficha_start = self._get_value('FICHA', 'start_date')
        if not ficha_start:
            ficha_start = contrato_start if contrato_start else start
        ws[f'D{current_row}'] = ficha_start
        ws[f'H{current_row}'] = start
        current_row += 1
        
        # Finish date
        ws[f'B{current_row}'] = 'Finish date'

        finish = self._get_value('CERTIFICADO', 'finish_date')
        contrato_finish = self._get_value('CONTRATO', 'finish_date')
        if not contrato_finish:
            contrato_finish = finish
        ws[f'C{current_row}'] = contrato_finish
        ficha_finish = self._get_value('FICHA', 'finish_date')
        if not ficha_finish:
            ficha_finish = contrato_finish if contrato_finish else finish
        ws[f'D{current_row}'] = ficha_finish
        ws[f'H{current_row}'] = finish
        current_row += 1
        
        # Address (ACT)
        ws[f'B{current_row}'] = 'Address'

        act_contr = self._get_value('CONTRATO', 'location')
        act_fact  = self._get_value('FACTURA', 'homeowner_address')
        act_cert  = self._get_value('CERTIFICADO', 'address')
        act_decl = self._clean_ocr_text(self._get_value('DECLARACION', 'homeowner_address'))

        # Depuración en español para ver los valores de dirección
        print(f"[DEPURACIÓN] Dirección ACT - CONTRATO: '{act_contr}' | DECLARACION: '{act_decl}' | FACTURA: '{act_fact}'")

        ws[f'C{current_row}'] = act_contr
        ws[f'E{current_row}'] = act_decl
        ws[f'F{current_row}'] = act_fact
        ws[f'H{current_row}'] = act_cert
        ws[f'I{current_row}'] = self._get_value('CEE', 'address')
        ws[f'J{current_row}'] = self._get_value('REGISTRO', 'address')
        ws[f'L{current_row}'] = ""  # no metas nombre aquí

        current_row += 1

        
        # Catastral ref
        ws[f'B{current_row}'] = 'Catastral ref'

        contr_cat = self._get_value('CONTRATO', 'catastral_ref')
        decl_cat  = self._get_value('DECLARACION', 'catastral_ref')
        cert_cat  = self._get_value('CERTIFICADO', 'catastral_ref')
        cee_cat   = self._get_value('CEE', 'catastral_ref')
        reg_cat   = self._get_value('REGISTRO', 'catastral_ref')

        # Si decl_cat es vacío o NOT FOUND, usa contr_cat
        decl_cat_excel = decl_cat
        if not self._is_valid_cell_value(decl_cat, min_len=8):
            if self._is_valid_cell_value(contr_cat, min_len=8):
                decl_cat_excel = contr_cat

        ws[f'C{current_row}'] = contr_cat
        ws[f'E{current_row}'] = decl_cat_excel
        ws[f'H{current_row}'] = cert_cat
        ws[f'I{current_row}'] = cee_cat
        ws[f'J{current_row}'] = reg_cat
        current_row += 1

        
        # UTM coordinates
        ws[f'B{current_row}'] = 'UTM coordinates'
        ws[f'C{current_row}'] = self._get_value('CONTRATO', 'utm_coordinates')
        current_row += 1
        
        # Investment (€)
        ws[f'B{current_row}'] = 'Investment (€)'
        ws[f'C{current_row}'] = self._get_value('CONTRATO', 'sell_price')
        current_row += 1
        
        # Remuneration
        ws[f'B{current_row}'] = 'Remuneration'
        current_row += 1
        
        # Lifespan
        ws[f'B{current_row}'] = 'Lifespan (años)'
        life = self._get_value('CERTIFICADO', 'lifespan')
        ws[f'C{current_row}'] = self._get_value('CONTRATO', 'lifespan')
        ws[f'H{current_row}'] = life
        current_row += 1
        
        # Sell price
        ws[f'B{current_row}'] = 'Sell price (€/kWh)'
        ws[f'C{current_row}'] = self._get_value('CONTRATO', 'sell_price')
        ws[f'D{current_row}'] = self._get_value('FICHA', 'sell_price')
        current_row += 1
        
        # Fp
        ws[f'B{current_row}'] = 'Fp'
        ficha_fp = self._get_value('FICHA', 'fp') or self._get_value('CALCULO', 'fp')
        ws[f'D{current_row}'] = self._norm_num(ficha_fp)
        ws[f'H{current_row}'] = self._norm_num(self._get_value('CERTIFICADO', 'fp'))
        ws[f'L{current_row}'] = self._norm_num(self._get_value('CALCULO', 'fp'))
        current_row += 1
        
        # Ui (Ki)
        ws[f'B{current_row}'] = 'Ui'
        ficha_ui = self._get_value('FICHA', 'ui') or self._get_value('CALCULO', 'ki')
        ws[f'D{current_row}'] = self._norm_num(ficha_ui)
        ws[f'H{current_row}'] = self._norm_num(self._get_value('CERTIFICADO', 'ui'))
        ws[f'L{current_row}'] = self._norm_num(self._get_value('CALCULO', 'ki'))
        current_row += 1
        
        # Uf (Kf)
        ws[f'B{current_row}'] = 'Uf'
        ficha_uf = self._get_value('FICHA', 'uf') or self._get_value('CALCULO', 'kf')
        ws[f'D{current_row}'] = self._norm_num(ficha_uf)
        ws[f'H{current_row}'] = self._norm_num(self._get_value('CERTIFICADO', 'uf'))
        ws[f'L{current_row}'] = self._norm_num(self._get_value('CALCULO', 'kf'))
        current_row += 1
        
        # S (Surface)
        ws[f'B{current_row}'] = 'S'
        ficha_s = self._get_value('FICHA', 'surface') or self._get_value('CALCULO', 's')
        ws[f'D{current_row}'] = self._norm_num(ficha_s)
        ws[f'E{current_row}'] = self._norm_num(self._get_value('DECLARACION', 'surface'))
        ws[f'F{current_row}'] = self._norm_num(self._get_value('FACTURA', 's'))
        ws[f'H{current_row}'] = self._norm_num(self._get_value('CERTIFICADO', 'surface'))
        ws[f'L{current_row}'] = self._norm_num(self._get_value('CALCULO', 's'))
        current_row += 1

        # b (surface coefficient / coef. zona)
        ws[f'B{current_row}'] = 'b'
        ws[f'H{current_row}'] = self._norm_num(self._get_value('CERTIFICADO', 'b'))
        ws[f'L{current_row}'] = self._norm_num(self._get_value('CALCULO', 'b'))
        current_row += 1

        # Climatic zone
        ws[f'B{current_row}'] = 'Climatic zone'
        ws[f'D{current_row}'] = self._get_value('FICHA', 'climatic_zone')
        ws[f'H{current_row}'] = self._get_value('CERTIFICADO', 'climatic_zone')  # E1
        ws[f'I{current_row}'] = self._get_value('CEE', 'climatic_zone')  # Agregar CEE
        ws[f'L{current_row}'] = ""  # OJO: 74 NO es zona, es G
        current_row += 1

        # G (surface coefficient / coef. zona) -> en tu caso 74
        ws[f'B{current_row}'] = 'G'
        ws[f'D{current_row}'] = self._norm_num(self._get_value('CERTIFICADO', 'g'))
        ws[f'H{current_row}'] = self._norm_num(self._get_value('CERTIFICADO', 'g'))  # 74,0
        ws[f'L{current_row}'] = self._norm_num(self._get_value('CALCULO', 'zone_climatique'))  # 74
        current_row += 1

        ws[f'B{current_row}'] = 'Type (ROLLO/SOPLADO)'
        ws[f'H{current_row}'] = self._get_value('CERTIFICADO', 'isolation_type')
        ws[f'F{current_row}'] = self._get_value('FACTURA', 'isolation_type')
        ws[f'D{current_row}'] = self._get_value('FICHA', 'isolation_type')
        ws[f'E{current_row}'] = self._get_value('DECLARACION', 'isolation_type')
        current_row += 1

        # Isolation Thickness
        ws[f'B{current_row}'] = 'Isolation Thickness'

        iso = self._get_value('CERTIFICADO', 'isolation_thickness')
        ws[f'C{current_row}'] = self._get_value('CONTRATO', 'isolation_thickness')
        ws[f'D{current_row}'] = self._get_value('FICHA', 'isolation_thickness')
        ws[f'F{current_row}'] = iso  # ✅ AÑADIR FACTURA
        ws[f'H{current_row}'] = iso
        current_row += 1
        
        # Surface coefficient (porcentaje del CALCULO)
        ws[f'B{current_row}'] = 'Surface coefficient'

        # Solo L: porcentaje de CALCULO
        calc_surf_coeff = self._get_value('CALCULO', 'porcentaje')
        ws[f'H{current_row}'] = ""  # CERTIFICADO no tiene surface coefficient separado
        ws[f'L{current_row}'] = calc_surf_coeff

        current_row += 1

        # Calculation methodology (R*t)
        ws[f'B{current_row}'] = 'Calculation methodology (R*t)'# Calculation methodology (R*t) - SOLO el R*t del CERTIFICADO

        rt = self._get_value('CERTIFICADO', 'calculation_methodology')  # 0,75

        ws[f'H{current_row}'] = self._norm_num(rt)
        ws[f'L{current_row}'] = ""  # NO poner 0,83 aquí

        current_row += 1
        # Photos (firma del Cedente desde el CONTRATO)
        ws[f'B{current_row}'] = 'Photos'

        # Dale altura a la fila para que quepa la imagen
        ws.row_dimensions[current_row].height = 110

        # ✅ CAMBIO 1: Limpiar las celdas ANTES de insertar imágenes
        ws[f'C{current_row}'] = None
        ws[f'E{current_row}'] = None

        # ✅ CAMBIO 2: Crear directorio temporal
        tmp_dir = Path(output_path).parent / "_tmp_assets"
        tmp_dir.mkdir(parents=True, exist_ok=True)

        # ✅ CAMBIO 3: Extraer e insertar firma del CONTRATO (columna C)
        contrato_pdf = str(self.file_mapping.get("CONTRATO", ""))
        if contrato_pdf.lower().endswith(".pdf") and os.path.exists(contrato_pdf):
            img_path = str(tmp_dir / "contrato_firma_cedente.png")
            try:
                self._extract_cedente_signature(contrato_pdf, img_path)
                # Verificar que la imagen se creó correctamente
                if os.path.exists(img_path) and os.path.getsize(img_path) > 0:
                    self._insert_image(ws, f"C{current_row}", img_path, width_px=260)
                    print(f"   ✅ Firma CONTRATO insertada en C{current_row}")
                    # ❌ NO PONGAS: ws[f'C{current_row}'] = ""
                else:
                    ws[f'C{current_row}'] = "Signature image empty"
            except Exception as e:
                ws[f'C{current_row}'] = f"Signature error: {str(e)[:40]}"
                print(f"   ❌ Error firma CONTRATO: {e}")

            # ✅ CAMBIO 4: Extraer e insertar firma de DECLARACION (columna E)
            declaracion_pdf = str(self.file_mapping.get("DECLARACION", ""))
            if declaracion_pdf.lower().endswith(".pdf") and os.path.exists(declaracion_pdf):
                img_path_decl = str(tmp_dir / "declaracion_firma_cedente.png")
                try:
                    self._extract_declaracion_signature(declaracion_pdf, img_path_decl)
                    # Verificar que la imagen se creó correctamente
                    if os.path.exists(img_path_decl) and os.path.getsize(img_path_decl) > 0:
                        self._insert_image(ws, f"E{current_row}", img_path_decl, width_px=260)
                        print(f"   ✅ Firma DECLARACION insertada en E{current_row}")
                        # ❌ NO PONGAS: ws[f'E{current_row}'] = ""
                    else:
                        # Fallback: intentar usar la firma del CONTRATO
                        img_path_contrato = str(tmp_dir / "contrato_firma_cedente.png")
                        if os.path.exists(img_path_contrato) and os.path.getsize(img_path_contrato) > 0:
                            self._insert_image(ws, f"E{current_row}", img_path_contrato, width_px=260)
                            print(f"   ✅ Firma CONTRATO reutilizada en E{current_row}")
                        else:
                            ws[f'E{current_row}'] = "Signature not found"
                except Exception as e:
                    ws[f'E{current_row}'] = f"Signature error: {str(e)[:40]}"
                    print(f"   ❌ Error firma DECLARACION: {e}")

        current_row += 1
                
        # ==================== INSTALLER SECTION ====================
        ws[f'A{current_row}'] = 'INSTALLER'
        ws[f'A{current_row}'].font = section_font
        
        # Installer Name
        ws[f'B{current_row}'] = 'Name'
        ws[f'C{current_row}'] = self._get_value('CONTRATO', 'installer')
        ws[f'F{current_row}'] = self._get_value('FACTURA', 'installer_name')
        ws[f'H{current_row}'] = self._get_value('CERTIFICADO', 'installer_name')
        current_row += 1
        
        # Installer Address
        ws[f'B{current_row}'] = 'Address'
        ws[f'C{current_row}'] = self._get_value('CONTRATO', 'installer_address')
        ws[f'F{current_row}'] = self._get_value('FACTURA', 'installer_address')
        current_row += 1
        
        # Installer CIF
        ws[f'B{current_row}'] = 'CIF'
        ws[f'C{current_row}'] = self._get_value('CONTRATO', 'installer_cif')
        ws[f'F{current_row}'] = self._get_value('FACTURA', 'installer_cif')
        current_row += 1
        
        # ==================== BILL SECTION ====================
        ws[f'A{current_row}'] = 'BILL'
        ws[f'A{current_row}'].font = section_font
        
        # Invoice Number
        ws[f'B{current_row}'] = 'Number'
        ws[f'F{current_row}'] = self._get_value('FACTURA', 'invoice_number')
        current_row += 1
        
        # Invoice Date
        ws[f'B{current_row}'] = 'Date'
        ws[f'F{current_row}'] = self._get_value('FACTURA', 'invoice_date')
        current_row += 1
        
        # Subtotal
        ws[f'B{current_row}'] = 'Subtotal'
        ws[f'F{current_row}'] = self._get_value('FACTURA', 'subtotal')
        current_row += 1
        
        # Deduction
        ws[f'B{current_row}'] = 'Deduction (CAE)'
        ws[f'F{current_row}'] = self._get_value('FACTURA', 'deduction')
        current_row += 1
        
        # A pagar
        ws[f'B{current_row}'] = 'A pagar'
        ws[f'F{current_row}'] = self._get_value('FACTURA', 'amount')
        current_row += 1
        
        # ==================== REGISTRO SECTION ====================
        ws[f'A{current_row}'] = 'REGISTRO'
        ws[f'A{current_row}'].font = section_font
        
        # Registration Number
        ws[f'B{current_row}'] = 'Number'
        ws[f'J{current_row}'] = self._get_value('REGISTRO', 'registration_number')
        current_row += 1
        
        # Registration Date
        ws[f'B{current_row}'] = 'Date'
        ws[f'J{current_row}'] = self._get_value('REGISTRO', 'registration_date')
        current_row += 1
        
        # ==================== CEE SECTION ====================
        ws[f'A{current_row}'] = 'CEE'
        ws[f'A{current_row}'].font = section_font
        
        # Certification Date
        ws[f'B{current_row}'] = 'Certification date'
        ws[f'I{current_row}'] = self._get_value('CEE', 'certification_date')
        current_row += 1
        
        # Estado
        ws[f'B{current_row}'] = 'Estado'
        ws[f'L{current_row}'] = self._get_value('CALCULO', 'estado')
        current_row += 1
        
        # Save
        wb.save(output_path)
        print(f"\n✅ Excel generated: {output_path}")
    
    def _get_value(self, doc_type: str, field: str) -> str:
        """Get value from parsed data"""
        if doc_type not in self.parsed_data:
            return ""

        value = self.parsed_data[doc_type].get(field, "")

        if value == "NOT FOUND" or value is None:
            return ""

        return str(value)

    def _insert_image(self, ws, cell: str, img_path: str, width_px: int = 280):
        img = XLImage(img_path)
        if img.width:
            ratio = width_px / img.width
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
        ws.add_image(img, cell)

    def _norm_num(self, v: Any) -> str:
        """
        Normaliza números para Excel:
        - 46.0 -> 46
        - 1,0 -> 1
        - 10314.00 -> 10314
        Mantiene decimales reales (3,26 se queda 3,26)
        """
        if v is None:
            return ""
        s = str(v).strip()
        if not s or s.upper() == "NOT FOUND":
            return ""

        # cambia punto a coma (pero si ya viene con coma, ok)
        s = s.replace(".", ",")

        # si es número con ,000... lo deja sin parte decimal
        if re.fullmatch(r"\d+,0+", s):
            return s.split(",")[0]

        # si es número con decimales tipo 10314,00 -> 10314
        if re.fullmatch(r"\d+,\d+", s):
            entero, dec = s.split(",", 1)
            if set(dec) == {"0"}:
                return entero

        return s

    def _is_valid_cell_value(self, v: Any, min_len: int = 2) -> bool:
        if v is None:
            return False
        s = str(v).strip()
        if not s:
            return False
        if s.upper() == "NOT FOUND":
            return False
        if len(s) < min_len:
            return False

        # --- NEW: clean OCR garbage prefixes like "! or similar ---
        if re.match(r'^[""\'`´!¡¿\W]{1,3}\s*[A-ZÁÉÍÓÚÑ]', s):
            print(f"🐛 REJECTED by prefix filter: {s}")
            return False
        
        if s[:2] in {"“!", "\"!", "’!", "‘!"}:
            print(f"🐛 REJECTED by combo filter: {s}")
            return False

        # basura típica OCR
        if s.upper() in {"C", "CL", "CALLE", "AV", "S/N"}:
            return False
        # evita MRZ basura típica del DNI
        # evita MRZ basura típica del DNI (pero NO catastrales válidos)
        if re.fullmatch(r"[A-Z0-9<]{10,}", s.upper()):
            # ✅ Excepción: Si tiene 14+ chars Y mezcla números/letras, probablemente es catastral
            if len(s) >= 14 and re.search(r'\d', s) and re.search(r'[A-Z]', s.upper()):
                pass  # Es catastral válido, no rechazar
            else:
                return False
        if s.upper().endswith(" MANEC"):
            return False
        return True


    def _pick_best(self, *candidates: Any, min_len: int = 2) -> str:
        for c in candidates:
            if self._is_valid_cell_value(c, min_len=min_len):
                return str(c).strip()
        return ""

    def _pick_best_energy_savings(self, contr_ae: str, cert_ae: str, calc_ae: str) -> str:
        """Pick best energy savings value with intelligent fallback logic."""
        # Helper to check if value is valid and convert to float
        def is_valid_energy(value: str) -> tuple[bool, float]:
            if not self._is_valid_cell_value(value, min_len=2):
                return False, 0.0
            try:
                num = float(value.replace(",", "."))
                return True, num
            except ValueError:
                return False, 0.0

        contr_valid, contr_num = is_valid_energy(contr_ae)
        cert_valid, cert_num = is_valid_energy(cert_ae)
        calc_valid, calc_num = is_valid_energy(calc_ae)

        # Priority: CERTIFICADO > CALCULO > CONTRATO (but avoid small contract values)
        if cert_valid:
            return cert_ae
        elif calc_valid:
            return calc_ae
        elif contr_valid and contr_num >= 500:  # Only use contract if it's reasonably large
            return contr_ae
        else:
            return ""

    def to_010_020(self, act_code: str) -> str:
        if not act_code or act_code == "NOT FOUND":
            return ""
        m = re.search(r"(010|020)", str(act_code))
        return m.group(1) if m else ""

    def _extract_cedente_signature(self, pdf_path: str, out_path: str) -> str:
        """
        Extrae la firma del cedente del CONTRATO.
        ✅ Siempre usa la última página del PDF.
        """
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        
        # Siempre usar la última página
        page = doc[-1]
        
        w, h = page.rect.width, page.rect.height
        # Cuadrado inferior izquierdo: desde 60% abajo, más arriba
        x0 = 0
        x1 = w * 0.5
        y0 = h * 0.6
        y1 = h * 0.9
        
        clip = fitz.Rect(x0, y0, x1, y1)
        
        os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
        zoom = 2  # Reducir zoom para mejor calidad
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=clip, alpha=False)
        pix.save(out_path)
        
        print(f"   📝 Firma extraída de CONTRATO página {total_pages}/{total_pages}, área: x0={x0:.0f}, y0={y0:.0f}, x1={x1:.0f}, y1={y1:.0f}")
        doc.close()
        return out_path

    def _extract_declaracion_signature(self, pdf_path: str, out_path: str) -> str:
        """
        Extrae firma de DECLARACIÓN (última página del documento).
        """
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        page = doc[-1]  # última página para DECLARACIÓN

        w, h = page.rect.width, page.rect.height

        # Clip ajustado: mover a la derecha, más espacio derecha, cortar arriba
        x0 = w * 0.55
        x1 = w
        y0 = h * 0.65
        y1 = h * 0.9

        clip = fitz.Rect(x0, y0, x1, y1)

        os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
        zoom = 2  # Igual que CONTRATO para consistencia
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=clip, alpha=False)
        pix.save(out_path)
        
        print(f"   📝 Firma extraída de DECLARACION página {total_pages}/{total_pages}")
        doc.close()
        return out_path

    def _clean_ocr_text(self, v: Any) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        if not s or s.upper() == "NOT FOUND":
            return ""

        # normaliza comillas raras
        s = s.replace("“", "").replace("”", "").replace("’", "").replace("‘", "")

        # quita basura OCR al principio tipo "! , . ; : etc"
        s = re.sub(r'^[\s"\'`´!¡¿\W]{1,5}', '', s)

        return s.strip()


