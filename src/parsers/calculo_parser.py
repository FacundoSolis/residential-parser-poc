"""
Parser for CALCULO.xlsx files
"""

import openpyxl
from pathlib import Path
from typing import Dict, Any


class CalculoParser:
    """Parser for Calculo Excel files"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.data = {}
        
    def parse(self) -> Dict[str, Any]:
        """Extract calculation data from Excel"""
        if not Path(self.file_path).exists():
            return {
                'document_type': 'CALCULO',
                'error': 'File not found'
            }
        
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            result = {
                'document_type': 'CALCULO',
                'client_name': self._extract_client_name(ws),
                'area_total': self._extract_area_total(ws),
                'area_afect': self._extract_area_afect(ws),
                'porcentaje': self._extract_porcentaje(ws),
                'act_code': self._extract_act_code(ws),
                'fp': self._extract_fp(ws),
                'ki': self._extract_ki(ws),
                'kf': self._extract_kf(ws),
                's': self._extract_s(ws),
                'zone_climatique': self._extract_zone(ws),
                'ae': self._extract_ae(ws),
                'estado': self._extract_estado(ws),
                'calculation_methodology': self._extract_methodology(ws),
            }
            
            return result
        except Exception as e:
            return {
                'document_type': 'CALCULO',
                'error': str(e)
            }
    
    def _extract_client_name(self, ws) -> str:
        """Extract client name from Row 1"""
        try:
            cell = ws['A1'].value
            if cell and 'Client' in str(cell):
                parts = str(cell).split(':')
                if len(parts) > 1:
                    name = parts[1].split('(')[0].strip()
                    return name
        except:
            pass
        return "NOT FOUND"
    
    def _extract_area_total(self, ws) -> str:
        """Extract total area"""
        try:
            for row in ws.iter_rows(min_row=3, max_row=5, values_only=True):
                if len(row) > 15:
                    if row[14] == 'Área total' and row[15] is not None:
                        return self._format_num(row[15], nd=2)

        except:
            pass
        return "NOT FOUND"
    
    def _extract_area_afect(self, ws) -> str:
        """Extract affected area"""
        try:
            for row in ws.iter_rows(min_row=3, max_row=5, values_only=True):
                if len(row) > 15:
                    if row[14] == 'Área afect' and row[15] is not None:
                        return str(row[15])
        except:
            pass
        return "NOT FOUND"

    def _format_percent(self, v) -> str:
        if v is None:
            return "NOT FOUND"
        try:
            x = float(v)
            if x <= 0:
                return "NOT FOUND"
            if x < 1:       # 0.1689 -> 16.89
                x *= 100
            return f"{x:.2f}%"
        except:
            s = str(v).strip()
            return s if s else "NOT FOUND"

    def _format_decimal_comma(self, v) -> str:
        if v is None:
            return "NOT FOUND"
        s = str(v).strip()
        if not s:
            return "NOT FOUND"
        return s.replace(".", ",")

    
    def _extract_porcentaje(self, ws) -> str:
        """Extract percentage"""
        try:
            for row in ws.iter_rows(min_row=5, max_row=6, values_only=True):
                if len(row) > 15 and row[14] == 'Porcentaje' and row[15] is not None:
                    return self._format_percent(row[15])
        except:
            pass
        return "NOT FOUND"
    def _format_percent(self, v) -> str:
        if v is None:
            return "NOT FOUND"
        try:
            x = float(v)
            if x <= 0:
                return "NOT FOUND"
            if x < 1:
                x *= 100
            return f"{x:.2f}%"
        except:
            return "NOT FOUND"


    
    def _extract_act_code(self, ws) -> str:
        """Extract RES code - look for RES020 or RES010 in column O (index 14)"""
        try:
            for row in ws.iter_rows(min_row=10, max_row=15, values_only=True):
                if len(row) > 14 and row[14] is not None:
                    val = str(row[14])
                    if 'RES' in val:
                        return val
        except:
            pass
        return "NOT FOUND"
    
    def _extract_fp(self, ws) -> str:
        """Extract Fp value - same row as RES, column P (index 15)"""
        try:
            for row in ws.iter_rows(min_row=10, max_row=15, values_only=True):
                if len(row) > 15 and row[14] is not None and 'RES' in str(row[14]):
                    if row[15] is not None:
                        return str(row[15])
        except:
            pass
        return "NOT FOUND"
    
    def _extract_ki(self, ws) -> str:
        """Extract Ki value - same row as RES, column Q (index 16)"""
        try:
            for row in ws.iter_rows(min_row=10, max_row=15, values_only=True):
                if len(row) > 16 and row[14] is not None and 'RES' in str(row[14]):
                    if row[16] is not None:
                        return str(row[16])
        except:
            pass
        return "NOT FOUND"
    
    def _extract_kf(self, ws) -> str:
        """Extract Kf value - same row as RES, column R (index 17)"""
        try:
            for row in ws.iter_rows(min_row=10, max_row=15, values_only=True):
                if len(row) > 17 and row[14] is not None and 'RES' in str(row[14]):
                    if row[17] is not None:
                        return str(row[17])
        except:
            pass
        return "NOT FOUND"
    
    def _extract_s(self, ws) -> str:
        """Extract S (surface) value - same row as RES, column S (index 18)"""
        try:
            for row in ws.iter_rows(min_row=10, max_row=15, values_only=True):
                if len(row) > 18 and row[14] is not None and 'RES' in str(row[14]):
                    if row[18] is not None:
                        return self._format_num(row[18], nd=2)

        except:
            pass
        return "NOT FOUND"
    
    def _extract_zone(self, ws) -> str:
        """Extract climatic zone - same row as RES, column T (index 19)"""
        try:
            for row in ws.iter_rows(min_row=10, max_row=15, values_only=True):
                if len(row) > 19 and row[14] is not None and 'RES' in str(row[14]):
                    if row[19] is not None:
                        return str(row[19])
        except:
            pass
        return "NOT FOUND"
    
    def _extract_ae(self, ws) -> str:
        """Extract AE (energy savings) - same row as RES, column U (index 20)"""
        try:
            for row in ws.iter_rows(min_row=10, max_row=15, values_only=True):
                if len(row) > 20 and row[14] is not None and 'RES' in str(row[14]):
                   if row[20] is not None:
                        return self._format_kwh(row[20])

        except:
            pass
        return "NOT FOUND"
    
    def _extract_estado(self, ws) -> str:
        """Extract ESTADO value - Row 16, column R (index 17)"""
        try:
            for row in ws.iter_rows(min_row=15, max_row=18, values_only=True):
                if len(row) > 17:
                    if row[14] == 'ESTADO:' and row[17] is not None:
                        return str(row[17])
        except:
            pass
        return "NOT FOUND"
    
    def _extract_methodology(self, ws) -> str:
        """Extract calculation methodology value"""
        try:
            for row_idx, row in enumerate(ws.iter_rows(min_row=16, max_row=20, values_only=True), 16):
                if len(row) > 17 and row[14] is not None and row[17] is not None:
                    desc = str(row[14]).lower()

                    if 'puertas' in desc or 'ventanas' in desc or 'aberturas' in desc:
                        val = row[17]
                        if val is not None:
                            return self._format_decimal_comma(val)

                    elif row[14] != 'ESTADO:':
                        val = row[17]
                        if isinstance(val, (int, float, str)):
                            return self._format_decimal_comma(val)
        except:
            pass
        return "NOT FOUND"

    def _format_num(self, v, nd=2) -> str:
        if v is None:
            return "NOT FOUND"
        try:
            x = float(v)
            # si es casi entero, devuelve entero
            if abs(x - round(x)) < 0.01:
                return str(int(round(x)))
            s = f"{x:.{nd}f}".rstrip("0").rstrip(".")
            return s
        except:
            # si ya viene como string bien, lo devuelves
            return str(v).strip() if str(v).strip() else "NOT FOUND"

    def _format_kwh(self, v) -> str:
        if v is None:
            return "NOT FOUND"
        try:
            x = float(v)
            if abs(x - round(x)) < 0.01:
                return str(int(round(x)))
            return f"{x:.2f}".rstrip("0").rstrip(".")
        except:
            return str(v).strip() if str(v).strip() else "NOT FOUND"

def _format_percent(self, v) -> str:
    if v is None:
        return "NOT FOUND"
    try:
        x = float(v)
        if x <= 0:
            return "NOT FOUND"
        if x < 1:
            x *= 100
        return f"{x:.2f}%"
    except:
        # si viene "16.89%" o "0,83" etc
        s = str(v).strip()
        return s if s else "NOT FOUND"

def _format_decimal_comma(self, v) -> str:
    if v is None:
        return "NOT FOUND"
    s = str(v).strip()
    if not s:
        return "NOT FOUND"
    return s.replace(".", ",")
